#!/usr/bin/python
# coding=UTF-8
"""
功能说明：        
        根据xlsx文件，生产多个场次的xml文件
        {
            多个sheet里面的数据项，会根据 level字段 分配到对应的 business_x.xml
            一个sheet可支持一种数组结构, 数组格式参考 ArraySample.xlsx
        }
        写入mysql，需要xlsx 的 mysql sheet里面填写正确的idx
        http post, 触发后台更新配置

        后续新游戏，根据旧game_business.xml模板，修改py代码调整xml对象树结构

命令格式: python excel_2_xml.py

g_cfg为全局配置，不同游戏自己修改， 里面有字段说明。

模块安装命令参考：
    window:
        http://www.codegood.com/downloads
"""
import hashlib
import requests
import pymysql as MySQLdb
import sys
import json
import openpyxl
import xml.etree.ElementTree as ET
reload(sys)
sys.setdefaultencoding('utf-8') 


def BuildArrayInfo(ws, g_cfg, array_info):   
    """
    根据重复字段，构建结构体数组结构体信息
    要求列顺序， 数组元素之间不能有其他非数组元素夹在中间。
    array_info [out]  in=[] . 比如。 10，19 列 为结构体和 20，29列为结构体， 组成2长度数值 array_info=[[10,19],[20,29]]  
    """
    m={} #key = field_name, value = col
    array_len=0
    array_info_idx=0
    first_field_name=0
    row = g_cfg['field_name_row']
    for col in range(1,ws.max_column+1):
        field_name = ws.cell(row,col).value
        if not field_name:
            continue
        #print field_name
        if 0 == first_field_name:#init
            if m.has_key(field_name):
                if len(array_info) == 0:
                    #print "first init first_field_name =" , field_name
                    #print row
                    array_info.append([])
                    array_info[array_info_idx].append(m[field_name])
                    array_len = col - m[field_name]
                    array_info[array_info_idx].append(m[field_name]+array_len-1)
                    array_info_idx = array_info_idx+1
                    first_field_name = field_name
            m[field_name]=col
        
        if first_field_name == field_name: #第一次成立，已经是第二个元素位置了
            array_info.append([])
            array_info[array_info_idx].append(col)
            array_info[array_info_idx].append(col+array_len-1)
            array_info_idx = array_info_idx+1
        

def IsArrayCol(array_info, col, dict):
    """retrun true表示是数组项，返回当前列所在的struct，在数组的索引"""
    for idx in range(len(array_info)):
        if col >= array_info[idx][0] and col <= array_info[idx][1]:
            dict['idx']=idx
            return True
    return False
    
def Text2Value(data_type, text):
    if data_type == "string":
        return text
    elif data_type == "double":
        try:
            return float(text)
        except:
            return 0.0 
    else:#未定义情况都认为是int
        try:
            return int(text)
        except:
            return 0 

def BuildExcelObj():
    """
    生成完成excel对象
    同时生成字段描述,层次结构和excel对象一样，最终值统一为字符串
    """
    excel_obj={}
    desc_obj={}
    inwb = openpyxl.load_workbook(g_cfg["excel_file"],data_only=True)
    sheetnames = inwb.get_sheet_names()   
    for sheet_name in sheetnames:
        ws = inwb.get_sheet_by_name(sheet_name)   #ws indicate one sheet
        rows=ws.max_row
        cols=ws.max_column

        excel_obj[sheet_name]={} #sheet 对应对象
        sheet_obj = excel_obj[sheet_name]
        desc_obj[sheet_name]={} #sheet 对应对象
        sheet_desc = desc_obj[sheet_name]
        array_info=[]
        BuildArrayInfo(ws, g_cfg, array_info) 

        #print "array_info=",array_info
        
        for row in range(g_cfg['start_row'],rows+1):
            #print "handle row:", row
            row_obj={}
            row_desc={}
            #if not ws.cell(row,1).value and ws.cell(row,1).value != 0:#如果第一列(可改)为空则不填充      
            #    continue
            if 'key_col' in g_cfg:
                key = str(ws.cell(row,g_cfg['key_col']).value)
            if not key is None: #表示无效值，或者缺省值
                key = row #默认key

            #print "key=", key    
            #init array
            array_len = len(array_info)
            array_name = "Array"+sheet_name
            if array_len > 0:
                #print "sheet, array_len=", sheet_name, array_len, array_info
                row_obj[array_name]=[]
                row_desc[array_name]=[]
                for col in range(0,array_len):
                    row_obj[array_name].append({})
                    row_desc[array_name].append({})

            for col in range(1,cols+1):
                ret={}
                data_type = ws.cell(g_cfg['type_row'],col).value # string or int
                if IsArrayCol(array_info, col, ret):#数组格式的其中一项
                    idx = ret['idx']
                    field_name = ws.cell(g_cfg['field_name_row'],col).value 
                    if not field_name:
                        continue
                    array_obj = row_obj[array_name]
                    array_desc = row_desc[array_name]

                    cell_v = ws.cell(g_cfg['desc_row'],col).value
                    if cell_v is None:
                        cell_v=""
                    if ws.cell(row,col).value:#空，没填的设为0
                        text=str(ws.cell(row,col).value)
                        array_obj[idx][field_name] = Text2Value(data_type, text)
                    else:
                        array_obj[idx][field_name] = Text2Value(data_type, "0") 
                    desc = str(ws.cell(g_cfg['desc_row'],col).value)
                    array_desc[idx][field_name] = desc
                            
                elif ws.cell(g_cfg['field_name_row'],col).value and  ws.cell(g_cfg['field_name_row'],col).value.strip()!= "":

                    if ws.cell(g_cfg['field_name_row'],col).value.find('/') != -1:#含有非法字符/
                        print("error,field name cannot contain '/' in row %d,col %d of %s(Sheet%d)"%(g_cfg['field_name_row'],col,excel_file,sheet_index+1))
                        exit(1)
                        
                    field_name = ws.cell(g_cfg['field_name_row'],col).value 

                    if ws.cell(row,col).value:#空，没填的设为0
                        text=str(ws.cell(row,col).value)
                        row_obj[field_name] = Text2Value(data_type, text)
                    else:
                        row_obj[field_name]= Text2Value(data_type, "") 
                    row_desc[field_name]= str(ws.cell(g_cfg['desc_row'],col).value)

            sheet_obj[key]=row_obj #end : for row in range(g_cfg['start_row'],rows+1):
            sheet_desc[key]=row_desc
    return excel_obj, desc_obj
    

def ToJsonStr(obj):
     str = json.dumps(obj, sort_keys=True, indent=4, separators=(',', ': '))
     return str
  
def WriteFile(file_name, str):
    fo = open(file_name, "w")
    fo.write( str )

def PrintJson(obj):
    None
    #print ToJsonStr(obj)

def ExcelObj2BusinessObj(excel_obj, desc_obj):
    """
    ExcelObj 结构，根据level划分为不同的business obj，和对应的 desc obj
    @excel_obj 
    @desc_obj 和excel_obj对应结构的描述对象
    """
    business_obj={}
    business_desc={}
    level_name = g_cfg["level"]#场次字段名

    #business_obj 构建多个场次空字典
    for sheet_name,sheet_obj in excel_obj.items():
        if g_cfg["level_sheet"] != sheet_name:
            continue

        for key,obj in sheet_obj.items():
            level = obj[level_name]
            if level: 
                None
            else:
                #print "error level=", level
                continue
            business_obj[level]={}
            business_desc[level]={}

    #business_obj 初始化 sheet_obj为[]
    for sheet_name,sheet_obj in excel_obj.items():
        for level,b_sheet_obj in business_obj.items():
            b_sheet_obj[sheet_name]=[]
        for level,b_sheet_obj in business_desc.items():
            b_sheet_obj[sheet_name]=[]
        
    #business_obj 构建每个数据
    for sheet_name,sheet_obj in excel_obj.items():
        for key,obj in sheet_obj.items():
            #print 'obj=', obj
            level = obj[level_name]
            level = int(level)
            b_key = str(key)
            if not b_key:
                b_key="None"
            if level: 
                business_obj[level][sheet_name].append(obj)
                business_desc[level][sheet_name].append(desc_obj[sheet_name][key])
            else:
                level = 0
                for b_level,b_sheet_obj in business_obj.items():
                    b_sheet_obj[sheet_name].append(obj)
                for b_level,b_sheet_obj in business_desc.items():
                    b_sheet_obj[sheet_name].append(desc_obj[sheet_name][key])


    return business_obj, business_desc

def ToXmlStr(xml_obj):
     str = ET.tostring(xml_obj, "UTF-8")
     return str



def BuildXmlFile(pre_file_name, business_obj, desc_obj):
    """
    @pre_file_name 路径文件前缀名，比如./out/business  ,表示生成 ./out/business_x.xml
    @business_obj python对象
    @desc_obj 对应的字段描述对象
    """
    for level,level_obj in business_obj.items():
        attrib = {"type":"struct",
                  "main":"true",
                  "name":g_cfg["head_name"],
                  "include":"base/xmlConfig",
                  "desc":"main class name and namespace"
                  }

        root_node = read_xml_file("",g_cfg["root_struct_name"],attrib)

      
        for key,sheet_value in level_obj.items():
            PythonObj2Xml(root_node, key, sheet_value, desc_obj[level][key]) 
            
        print "write to mysql"
        format_xml(root_node,'\t', '\n')
        xml_str= ET.tostring(root_node, encoding='UTF-8', method='xml')
        #WriteFile("out/t.xml", xml_str)
        UpdateToDb(level, xml_str)
        SendHttp(level)
        write_xml_file(root_node, pre_file_name + "_" + str(level) +".xml")

def WriteToDb(xml_str, cfg):
    """
    begin tran
    update table set  column=columnvalue   where wherestr
if @@ROWCOUNT==0
begin
    insert into table (column) values (columnvalue)
end 
commit tran

    """
    #cfg内容，以后配excel
    db_cfg={
            "1":1
        }


def format_xml(element, indent, newline, level = 0):
    if element is not None:
        if element.text == None or element.text.isspace(): # 如果element的text没有内容    
             element.text = newline + indent * (level + 1)      
        else:    
            #element.text = newline + indent * (level + 1) + element.text.strip() + newline + indent * (level + 1)    
            pass
    temp = list(element) # 将elemnt转成list    
    for subelement in temp:
        if temp.index(subelement) < (len(temp)-1):#如果不是list的最后一个元素,说明下一个行是同级别元素的起始,缩进应一致    
            subelement.tail = newline + indent * (level + 1)
            if level == 0:#TODO 额外加，二级之间空一行
                subelement.tail = newline + subelement.tail
        else:  # 如果是list的最后一个元素， 说明下一行是母元素的结束，缩进应该少一个    
            subelement.tail = newline + indent * level    

        format_xml(subelement, indent, newline, level = level + 1) # 对子元素进行递归操作  

def write_xml_file(root_node,xml_file):
	tree = ET.ElementTree(root_node)
	format_xml(root_node,'\t', '\n')
	tree.write(xml_file,"UTF-8")

def PythonObj2Xml(parent_xml_obj, obj_key, obj, desc_obj):
    """
    obj 转为xml对象，并加到 parent_xml_obj的节点
    """
    if type(obj) == type({}):
        sub_xml = ET.SubElement(parent_xml_obj, obj_key)#新建
        sub_xml.attrib = {"type":"struct"}
        for key,value in obj.items():
            PythonObj2Xml(sub_xml, key, value, desc_obj[key])
    elif type(obj) == type([]):
        sub_xml = ET.SubElement(parent_xml_obj, obj_key)
        sub_xml.attrib = {"type":"array", "value":"Item"+obj_key}
        for idx, value in enumerate(obj):
            PythonObj2Xml(sub_xml, 'Item'+obj_key, value, desc_obj[idx])
    elif type(obj) == type(1):
        sub_xml = ET.SubElement(parent_xml_obj, obj_key)
        sub_xml.attrib["type"]="int"
        sub_xml.attrib["desc"]=desc_obj
        sub_xml.text = str(obj)
    elif type(obj) == type(0.1):
        sub_xml = ET.SubElement(parent_xml_obj, obj_key)
        sub_xml.attrib["type"]="double"
        sub_xml.attrib["desc"]=desc_obj
        sub_xml.text = str(obj)
    elif type(obj) == type(str()):
        sub_xml = ET.SubElement(parent_xml_obj, obj_key)
        sub_xml.attrib["type"]="string"
        sub_xml.attrib["desc"]=desc_obj
        sub_xml.text = str(obj)
    else:
        print 'error, unknow type=', type(obj)
    return

def read_xml_file(xml_file,root_name,attrib={}):
    try:
        tree = ET.parse(xml_file)
        root_node = tree.getroot()
    except:
        root_node = ET.Element(root_name)
        root_node.attrib = attrib

    return root_node

def GetLevelMysqlCfg(level):
    """
    获取 mysql sheet 一行数据
    失败返回false
    """
    cfg = False #mysql sheet的一行配置
    for key,value in g_mysql_obj.items():
        if value["service_level"] == level:
            cfg = value
            break
    return cfg

def UpdateToDb(level, xml_str):
    
    #find cfg

    cfg = GetLevelMysqlCfg(level)
    if False == cfg:
        print "find mysql cfg fail in sheet. level=", level
        return

    # 打开数据库连接
    db = MySQLdb.connect(
       host=g_cfg["db_ip"],
       port=g_cfg["db_port"],
       user=g_cfg["db_user"],
       passwd=g_cfg["db_psw"],
       db=g_cfg["db_name"],
       charset="utf8"
       )

    # 使用cursor()方法获取操作游标 
    cursor = db.cursor()
    #cfg["service_id"]
    #service_id	service_level	service_name	service_path	business_file_name	service_file_name
    sql = """UPDATE tb_service_config_info SET 
    service_name = '%s',
    service_path = '%s',
    business_file_name = '%s',
    service_file_name = '%s',
    business_config_context = '%s' 
    WHERE idx = %d
    """ % ( MySQLdb.escape_string(cfg["service_name"]),
            MySQLdb.escape_string(cfg["service_path"]),
            MySQLdb.escape_string(cfg["business_file_name"]),
            MySQLdb.escape_string(cfg["service_file_name"]),
          MySQLdb.escape_string(xml_str) , cfg["idx"])
    
    #tb_service_config_info
    try:
       # 执行SQL语句
       cursor.execute(sql)
       # 提交到数据库执行
       db.commit()
       print "update db ok"
    except:
       # 发生错误时回滚
       db.rollback()
       print "update db fail"
    # 关闭数据库连接
    db.close()


def SeparateMysqlObj(excel_obj, desc_obj):
    """
    分离mysql sheet obj,从 excel_obj删除，并从返回mysql_obj
    """
    for sheet_name,sheet_obj in excel_obj.items():
        if sheet_name != "mysql":
            continue
        mysql_obj = sheet_obj
        del excel_obj[sheet_name]
        del desc_obj[sheet_name]
        return mysql_obj

def SendHttp(level):
    cfg = GetLevelMysqlCfg(level)
    if False == cfg:
        print "find mysql cfg fail in sheet. level=", level
        return

    url = 'http://'+g_cfg["http_ip"]+":"+str(g_cfg["http_port"])+"/update/config"
    content="file_context_type=business_config_context&file_name=business.xml&idx=%d&server_id=%d" % (cfg["idx"], 1)
    md5value=content+"&"+g_cfg["http_key"]
    md5hash = hashlib.md5(md5value)
    md5 = md5hash.hexdigest()
    body = content+"&signature="+ md5.upper(); 

    print "url=",url
    print "body=", body
    response = requests.post(url, data = body)

    print response.text
    print response.status_code

if __name__ == '__main__':

    global g_cfg #全局配置
    g_cfg={
        "excel_file":"./test.xlsx",     #excel文件路径
        "level_sheet":"CommonCfg",      #决定多少场次的sheet
        "level":"level",                #场次字段名
        "desc_row":1 ,                  #第几行是说明的
        "field_name_row":2,             #第几行是字段名
        "type_row":3,                   #第几行是数据类型,填 int double string
        "key_col":1,                    #第几列是标识key
        "start_row":5,                   #第几行开始是需要读取的有用的内容

        #xml 头的信息
        "head_name":"HeadName",
        "root_struct_name":"RootCfg",

        #db cfg
        "db_ip":"192.168.1.243",
        "db_port":5306,
        "db_user":"app",
        "db_psw":"123456",
        "db_name":"operation_db_app",   #"only_test_db"

        #http cfg
        "http_ip":"192.168.1.243",
        "http_port":40001,
        "http_key":"hmdfetycfgh2099#x*337709!md",

        }
    g_is_db=True
    
    global g_mysql_obj
    excel_obj, excel_desc_obj = BuildExcelObj();
    g_mysql_obj = SeparateMysqlObj(excel_obj, excel_desc_obj)
    business_obj, desc_obj = ExcelObj2BusinessObj(excel_obj, excel_desc_obj);

    #临时调试，总体浏览business obj用
    #WriteFile("out/all_for_view.json", ToJsonStr(business_obj))
    #WriteFile("out/desc_obj.json", ToJsonStr(excel_desc_obj))
    #exit(1)
    BuildXmlFile("out/business", business_obj, desc_obj)
    
    print "build ok"


